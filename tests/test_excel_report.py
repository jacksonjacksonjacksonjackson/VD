"""
test_excel_report.py

Smoke + structural assertions for the Excel report v2 (cover + 6 sheets).
Verifies sheet set/order, the canonical Cover assumptions block, that the old
$15,000 hardcoded EV premium is gone, incentive net <= gross, and the scenario
comparison table.
"""

import os
import tempfile

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook

from data.models import (
    FleetVehicle, VehicleIdentification, FuelEconomyData, Fleet,
    ElectrificationAnalysis, ChargingAnalysis, EmissionsInventory,
)
from analysis.reports import ExcelReportGenerator, cover_aref, COVER_SHEET_NAME, _ASSUMPTION_ROW


def _vehicle(i, acf="B"):
    vin = "1FTAA" + str(i).zfill(12)
    v = FleetVehicle(vin=vin)
    v.vehicle_id = VehicleIdentification(
        vin=vin, year="2017", make="Ford", model="F-350",
        fuel_type="Diesel", gvwr="Class 6")
    v.fuel_economy = FuelEconomyData(combined_mpg=12)
    v.processing_success = True
    v.annual_mileage = 15000
    v.odometer = 90000
    v.match_confidence = 80
    v.custom_fields.update({
        "_acf_code": acf, "ACF Category": acf,
        "_ev_purchase_price": 90000, "_payback_years": 8.0,
        "Proposed EV Year": str(2028 + (i % 3)),
    })
    return v


def _full_analysis(vehicles):
    an = ElectrificationAnalysis(fleet_name="RCSD")
    an.gas_price = 3.5
    an.electricity_price = 0.13
    an.ev_efficiency = 0.3
    an.discount_rate = 5.0
    an.analysis_period = 12
    an.total_savings = 200000
    an.payback_period = 8.0
    an.co2_savings = 60
    an.fleet_cash_flows = [
        {"year": y, "gas_price": 3.5, "electricity_price": 0.13} for y in range(1, 13)]
    an.vehicle_results = {
        v.vin: {"display_name": f"veh {i}", "annual_mileage": 15000, "mpg": 12,
                "annual_fuel_savings": 3000, "total_fuel_savings": 36000,
                "total_co2_reduction": 10, "total_npv_savings": 20000}
        for i, v in enumerate(vehicles)}
    an.prioritized_vehicles = [v.vin for v in vehicles]
    return an


@pytest.fixture
def workbook_path():
    vehicles = [_vehicle(i, acf=("A" if i == 0 else "B")) for i in range(6)]
    fleet = Fleet(name="RCSD")
    fleet.vehicles = vehicles
    fleet.max_vehicles_per_year = 2

    an = _full_analysis(vehicles)
    ch = ChargingAnalysis(fleet_name="RCSD")
    ch.level2_chargers_needed = 8
    ch.dcfc_chargers_needed = 3
    ch.estimated_installation_cost = 200000
    em = EmissionsInventory(fleet_name="RCSD")
    em.total_emissions = 400
    em.by_fuel_type = {"Diesel": 400}

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    ok = ExcelReportGenerator(path).generate(
        fleet=fleet, analysis=an, charging=ch, emissions=em, state_code="CA")
    assert ok
    yield path
    os.remove(path)


EXPECTED_SHEETS = [
    "Cover & Methodology",
    "Vehicle Data",
    "Fleet Overview",
    "TCO & Financials",
    "Replacement & Capital Plan",
    "Emissions & Scenarios",
    "Infrastructure & Action Items",
]


def test_sheet_set_and_order(workbook_path):
    wb = load_workbook(workbook_path)
    assert wb.sheetnames == EXPECTED_SHEETS


def test_cover_has_canonical_assumptions(workbook_path):
    wb = load_workbook(workbook_path)
    cov = wb[COVER_SHEET_NAME]
    # Gas price seed lives at the mapped canonical row, column B.
    gas_row = _ASSUMPTION_ROW["gas_price"]
    assert cov.cell(gas_row, 2).value == pytest.approx(3.5)
    # analysis_period is the last assumption
    period_row = _ASSUMPTION_ROW["analysis_period"]
    assert cov.cell(period_row, 2).value == pytest.approx(12)


def test_tco_assumptions_mirror_cover(workbook_path):
    wb = load_workbook(workbook_path)
    tco = wb["TCO & Financials"]
    # B4 should be a formula pointing at the Cover gas-price cell.
    assert tco["B4"].value == "=" + cover_aref("gas_price")


def test_no_hardcoded_15000_premium():
    """The old Electrification payback used a flat `ev_premium = 15000`; ensure the
    code no longer hardcodes an EV premium for payback."""
    import analysis.reports as reports_mod
    src = open(reports_mod.__file__, encoding="utf-8").read()
    assert "ev_premium" not in src, "Old hardcoded EV-premium payback logic still present"


def test_per_vehicle_payback_uses_real_value(workbook_path):
    wb = load_workbook(workbook_path)
    tco = wb["TCO & Financials"]
    # Find the per-vehicle savings header, then check a payback cell == 8.0
    found_header = False
    for r in range(1, tco.max_row + 1):
        if tco.cell(r, 1).value == "Per-Vehicle Savings Detail":
            found_header = True
            # header row is r+1, data starts r+2
            assert tco.cell(r + 2, 7).value == pytest.approx(8.0)
            break
    assert found_header


def test_incentive_net_not_greater_than_gross(workbook_path):
    wb = load_workbook(workbook_path)
    tco = wb["TCO & Financials"]
    for r in range(1, tco.max_row + 1):
        if tco.cell(r, 1).value == "TOTAL":
            gross = tco.cell(r, 3).value
            net = tco.cell(r, 6).value
            if isinstance(gross, (int, float)) and isinstance(net, (int, float)):
                assert net <= gross
            break


def test_scenario_comparison_present(workbook_path):
    wb = load_workbook(workbook_path)
    es = wb["Emissions & Scenarios"]
    names = {es.cell(r, 1).value for r in range(1, es.max_row + 1)}
    assert "Scenario Comparison (time-based presets)" in names


def test_action_items_present(workbook_path):
    wb = load_workbook(workbook_path)
    ia = wb["Infrastructure & Action Items"]
    labels = {ia.cell(r, 1).value for r in range(1, ia.max_row + 1)}
    assert "Data Gaps & Action Items" in labels


def test_generates_with_minimal_data():
    """A bare fleet (no analysis/charging/emissions) still yields a valid file."""
    v = _vehicle(0)
    fleet = Fleet(name="RCSD")
    fleet.vehicles = [v]
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        assert ExcelReportGenerator(path).generate(fleet=fleet)
        wb = load_workbook(path)
        # Cover, Vehicle Data, Fleet Overview, Replacement, Infrastructure always present
        assert "Cover & Methodology" in wb.sheetnames
        assert "Infrastructure & Action Items" in wb.sheetnames
    finally:
        os.remove(path)
