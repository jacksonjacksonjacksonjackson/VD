"""
results_panel.py

Panel for displaying and interacting with processed vehicle data in the
Fleet Electrification Analyzer.
"""

import logging
import os
import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, List, Any, Optional, Callable, Tuple
import threading

from settings import (
    PRIMARY_HEX_1,
    PRIMARY_HEX_2, 
    PRIMARY_HEX_3,
    SECONDARY_HEX_1,
    COLUMN_NAME_MAP,
    DEFAULT_VISIBLE_COLUMNS
)
from utils import SimpleTooltip
from data.models import FleetVehicle
from ui.theme import Colors, Fonts, Spacing

# Set up module logger
logger = logging.getLogger(__name__)

class ResultsPanel(ttk.Frame):
    """
    Panel for displaying and interacting with processed vehicle data.
    Features filterable and sortable treeview with customizable columns.
    """
    
    def __init__(self, parent, data=None, visible_columns=None,
               on_selection_change=None, on_column_change=None,
               db_manager=None, on_acf_override=None):
        """
        Initialize the results panel.

        Args:
            parent: Parent widget
            data: Initial data to display (list of FleetVehicle objects)
            visible_columns: Initially visible columns (or None for defaults)
            on_selection_change: Callback when selection changes
            on_column_change: Callback when column visibility changes
            db_manager: VehicleDatabaseManager instance (optional).  When
                provided, enables "Save MPG to Database" from the context menu.
            on_acf_override: Callback(vehicle) fired after an ACF category
                override is applied, so the main window can sync other panels.
        """
        super().__init__(parent)

        # Store callbacks
        self.on_selection_change_callback = on_selection_change
        self.on_column_change_callback = on_column_change
        self.db_manager = db_manager
        self.on_acf_override_callback = on_acf_override

        # Fleet type for single-vehicle ACF recalculation — updated by main_window
        # after Run Full Analysis so ACF override → recalculate uses the correct deadlines.
        self._fleet_type = "hpf"
        
        # Initialize variables
        self.data = data or []
        self.visible_columns = visible_columns or DEFAULT_VISIBLE_COLUMNS
        self.all_columns_map = {}  # Maps column IDs to display names
        self.search_var = tk.StringVar()  # For search/filter
        self.search_fields_var = tk.StringVar(value="all")  # Fields to search in
        self.status_filter_var = tk.StringVar(value="all")  # Status filter
        self.quality_filter_var = tk.StringVar(value="all")  # Quality filter
        self.data_map = {}  # Maps treeview IIDs to data objects
        
        # Create UI components
        self._create_toolbar()
        self._create_summary()
        self._create_treeview()
        
        # Populate data
        self.populate_data()
        
        # Create context menu
        self._create_context_menu()
    
    def _create_toolbar(self):
        """Create the toolbar with search, filters, and export."""
        toolbar = ttk.Frame(self)
        toolbar.pack(fill=tk.X, padx=Spacing.MARGIN_ELEMENT,
                     pady=(Spacing.MARGIN_ELEMENT, Spacing.SM))

        # ── Left: Search ────────────────────────────────────────────
        search_frame = ttk.Frame(toolbar)
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT, padx=(0, Spacing.XS))

        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=25)
        search_entry.pack(side=tk.LEFT, padx=(0, Spacing.XS))
        SimpleTooltip(search_entry, "Search across all visible columns — updates as you type")
        self.search_var.trace_add("write", lambda *args: self._apply_filter())

        ttk.Label(search_frame, text="in:").pack(side=tk.LEFT, padx=(Spacing.XS, 2))
        search_fields = ttk.Combobox(
            search_frame,
            textvariable=self.search_fields_var,
            values=["all", "VIN", "Make", "Model", "Year", "Asset ID", "Department"],
            width=10, state="readonly"
        )
        search_fields.pack(side=tk.LEFT, padx=(0, Spacing.MD))
        search_fields.bind("<<ComboboxSelected>>", lambda e: self._apply_filter())

        # ── Middle: Filters (compact, no LabelFrame) ───────────────
        sep = ttk.Separator(search_frame, orient=tk.VERTICAL)
        sep.pack(side=tk.LEFT, fill=tk.Y, padx=Spacing.SM, pady=2)

        ttk.Label(search_frame, text="Status:").pack(side=tk.LEFT, padx=(0, 2))
        status_filter = ttk.Combobox(
            search_frame,
            textvariable=self.status_filter_var,
            values=["all", "successful", "failed"],
            width=10, state="readonly"
        )
        status_filter.pack(side=tk.LEFT, padx=(0, Spacing.SM))
        SimpleTooltip(status_filter, "Filter by processing status")
        status_filter.bind("<<ComboboxSelected>>", lambda e: self._apply_filter())

        ttk.Label(search_frame, text="Quality:").pack(side=tk.LEFT, padx=(0, 2))
        quality_filter = ttk.Combobox(
            search_frame,
            textvariable=self.quality_filter_var,
            values=["all", "high (80%+)", "medium (50-80%)", "low (<50%)"],
            width=13, state="readonly"
        )
        quality_filter.pack(side=tk.LEFT, padx=(0, Spacing.SM))
        SimpleTooltip(quality_filter, "Filter by data quality score")
        quality_filter.bind("<<ComboboxSelected>>", lambda e: self._apply_filter())

        # Reset link (small, next to filters)
        reset_btn = ttk.Button(
            search_frame, text="Reset",
            command=self._clear_filter, width=5
        )
        reset_btn.pack(side=tk.LEFT)
        SimpleTooltip(reset_btn, "Clear all search and filter criteria")

        # ── Right: Export ───────────────────────────────────────────
        export_btn = ttk.Button(
            toolbar,
            text="Export",
            command=self._export_dialog,
            style="Primary.TButton"
        )
        export_btn.pack(side=tk.RIGHT, padx=Spacing.SM)
        SimpleTooltip(export_btn, "Export fleet data to CSV or Excel")
    
    def _create_summary(self):
        """Create the processing summary bar."""
        self.summary_frame = ttk.Frame(self)
        self.summary_frame.pack(fill=tk.X, padx=10, pady=(0, 5))

        # Configure summary frame style
        style = ttk.Style()
        style.configure("Summary.TFrame", relief="solid", borderwidth=1)
        self.summary_frame.configure(style="Summary.TFrame")

        # Main summary label (left side)
        self.summary_label = ttk.Label(
            self.summary_frame,
            text="No data loaded",
            font=(Fonts.FAMILY_SANS, Fonts.SIZE_BODY, "bold"),
            foreground="#718096"
        )
        self.summary_label.pack(side=tk.LEFT, padx=10, pady=5)

        # Vehicle count badge (right side)
        self.count_label = ttk.Label(
            self.summary_frame,
            text="",
            font=(Fonts.FAMILY_SANS, Fonts.SIZE_SMALL),
            foreground=Colors.TEXT_TERTIARY
        )
        self.count_label.pack(side=tk.RIGHT, padx=10, pady=5)
    
    def _create_treeview(self):
        """Create the treeview for displaying vehicle data."""
        # Create frame for treeview and scrollbars
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Build all column mapping
        self._build_column_map()
        
        # Create scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Create treeview
        self.tree = ttk.Treeview(
            tree_frame,
            columns=self.visible_columns,
            show="headings",
            selectmode="extended",
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure scrollbars
        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)
        
        # Configure columns and headings with content-aware widths
        for col in self.visible_columns:
            display_name = self.all_columns_map.get(col, col)
            self.tree.heading(
                col,
                text=display_name,
                command=lambda c=col: self._sort_by_column(c)
            )

            # Content-aware column widths
            width, anchor = self._get_column_width_and_anchor(col, display_name)
            self.tree.column(col, width=width, minwidth=50, anchor=anchor)
        
        # Configure color tags for data quality indication
        self.tree.tag_configure("failed", background="#FFE4E1")  # Light red for failed
        self.tree.tag_configure("high_quality", background="#F0FFF0")  # Light green for high quality
        self.tree.tag_configure("medium_quality", background="#FFFACD")  # Light yellow for medium quality  
        self.tree.tag_configure("low_quality", background="#FFE4B5")  # Light orange for low quality
        
        # Bind events
        self.tree.bind("<<TreeviewSelect>>", self._on_selection_change)
        self.tree.bind("<Double-1>", self._show_details)
        self.tree.bind("<Button-3>", self._show_context_menu)
    
    def _create_context_menu(self):
        """Create context menu for the treeview."""
        self.context_menu = tk.Menu(self, tearoff=0)
        
        # Add menu items
        self.context_menu.add_command(
            label="View Details",
            command=self._show_details
        )
        
        self.context_menu.add_command(
            label="Copy Selected",
            command=self.copy_selection
        )
        
        self.context_menu.add_separator()
        
        self.context_menu.add_command(
            label="Analyze Selected",
            command=self._analyze_selected
        )
        
        self.context_menu.add_separator()
        
        self.context_menu.add_command(
            label="Select All",
            command=self._select_all
        )
        
        self.context_menu.add_command(
            label="Deselect All",
            command=self._deselect_all
        )

        self.context_menu.add_separator()

        self.context_menu.add_command(
            label="Save MPG to Database",
            command=self._save_mpg_to_database
        )

        self.context_menu.add_separator()

        self.context_menu.add_command(
            label="Override ACF Category...",
            command=self._override_acf_category,
        )

        # Bind right-click to show menu
        self.tree.bind("<Button-3>", self._show_context_menu)
    
    def _build_column_map(self):
        """Build the mapping of all possible columns to display names.

        Unions keys across all vehicles (not just the first) so that
        custom_fields columns like ACF Category are discovered even if
        the first vehicle's processing failed.  Caps the scan at 50
        vehicles for performance on very large fleets.
        """
        # Start with column map from settings
        self.all_columns_map = dict(COLUMN_NAME_MAP)

        # Add any missing columns based on data — sample multiple vehicles
        if self.data:
            sample_size = min(len(self.data), 50)
            for vehicle in self.data[:sample_size]:
                for key in vehicle.to_row_dict().keys():
                    if key not in self.all_columns_map:
                        self.all_columns_map[key] = key

    @staticmethod
    def _get_column_width_and_anchor(col_id: str, display_name: str) -> Tuple[int, str]:
        """
        Return (pixel_width, anchor) for a column based on its data type.
        Numeric columns are right-aligned; text columns are left-aligned.
        VIN gets extra width for its 17-character values.
        """
        # Column width presets based on expected content
        COLUMN_WIDTHS = {
            # Wide columns — long text values
            "VIN": 170,
            "BodyClass": 150,
            "GVWR": 160,
            "Commercial Summary": 180,
            "Processing Error": 200,
            "Assumed Vehicle (Text)": 180,
            # Medium columns — short text
            "Make": 100,
            "Model": 120,
            "FuelTypePrimary": 120,
            "Department": 120,
            "Location": 120,
            "Asset ID": 100,
            "Commercial Category": 130,
            "Vehicle Class": 110,
            # Narrow columns — short numeric / coded values
            "Year": 70,
            "MPG City": 80,
            "MPG Highway": 90,
            "MPG Combined": 100,
            "CO2 emissions": 90,
            "GVWR (lbs)": 95,
            "Data Quality": 90,
            "Processing Status": 110,
            "Odometer": 95,
            "Annual Mileage": 105,
            "Is Diesel": 75,
            "Is Commercial": 90,
            # Match quality & MPG provenance
            "Match Confidence": 110,
            "Fuel Type Mismatch": 180,
            "MPG Source": 150,
            "MPG Estimated": 100,
            "EPA Class Est. MPG": 170,
            # ACF compliance & electrification
            "ACF Category": 140,
            "ACF Detail": 220,
            "ACF Relevance": 180,
            "Proposed EV Year": 115,
            "EV Year Reason": 250,
            # EV equivalent matching
            "EV Equivalent": 180,
            "EV MSRP Range": 150,
            "EV EPA Range": 100,
            "EV Fit Score": 100,
        }

        # Right-align numeric columns
        NUMERIC_COLUMNS = {
            "Year", "MPG City", "MPG Highway", "MPG Combined",
            "CO2 emissions", "co2A", "GVWR (lbs)", "Odometer",
            "Annual Mileage", "rangeA", "Engine HP",
            "Proposed EV Year", "Match Confidence",
        }

        width = COLUMN_WIDTHS.get(col_id, max(100, len(display_name) * 9))
        anchor = "e" if col_id in NUMERIC_COLUMNS else "w"
        return width, anchor

    def _sort_by_column(self, column):
        """
        Sort treeview data by specified column.
        
        Args:
            column: Column identifier to sort by
        """
        # Get current sort direction
        if hasattr(self, '_sort_column') and self._sort_column == column:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_column = column
            self._sort_reverse = False
        
        # Update heading text to show sort direction
        heading_text = self.all_columns_map.get(column, column)
        if self._sort_reverse:
            heading_text += " ▼"
        else:
            heading_text += " ▲"
        
        # Update heading
        self.tree.heading(column, text=heading_text)
        
        # Get all items
        items = [(self.tree.set(iid, column), iid) for iid in self.tree.get_children("")]
        
        # Sort items
        items.sort(reverse=self._sort_reverse)
        
        # Try to convert to numeric for numeric sorting
        def try_numeric_sort():
            try:
                # Convert to numeric if possible
                numeric_items = [(float(value), iid) for value, iid in items]
                return sorted(numeric_items, reverse=self._sort_reverse)
            except ValueError:
                # Fall back to string sort
                return None
        
        # Try numeric sort first
        numeric_sort = try_numeric_sort()
        if numeric_sort is not None:
            items = numeric_sort
        
        # Rearrange items
        for idx, (_, iid) in enumerate(items):
            self.tree.move(iid, "", idx)
        
        # Restore default headings for other columns
        for col in self.visible_columns:
            if col != column:
                self.tree.heading(col, text=self.all_columns_map.get(col, col))
    
    def _apply_filter(self):
        """Apply search and filter criteria to the treeview."""
        # Get current filter criteria
        search_text = self.search_var.get().strip().lower()
        search_field = self.search_fields_var.get()
        status_filter = self.status_filter_var.get()
        quality_filter = self.quality_filter_var.get()
        
        # Clear the treeview
        self.tree.delete(*self.tree.get_children())
        self.data_map.clear()
        
        # If no data, return
        if not self.data:
            return
        
        # Filter and display matching vehicles
        filtered_count = 0
        for vehicle in self.data:
            # Apply status filter
            if status_filter != "all":
                if status_filter == "successful" and not vehicle.processing_success:
                    continue
                if status_filter == "failed" and vehicle.processing_success:
                    continue
            
            # Apply quality filter
            if quality_filter != "all":
                quality_score = getattr(vehicle, 'data_quality_score', 0)
                if quality_filter.startswith("high") and quality_score < 80:
                    continue
                if quality_filter.startswith("medium") and (quality_score < 50 or quality_score >= 80):
                    continue
                if quality_filter.startswith("low") and quality_score >= 50:
                    continue
            
            # Apply search filter
            if search_text:
                match_found = False
                row_dict = vehicle.to_row_dict()
                
                if search_field == "all":
                    # Search all visible fields
                    for col in self.visible_columns:
                        value = str(row_dict.get(col, "")).lower()
                        if search_text in value:
                            match_found = True
                            break
                else:
                    # Search specific field
                    if search_field in row_dict:
                        value = str(row_dict[search_field]).lower()
                        match_found = search_text in value
                
                if not match_found:
                    continue
            
            # Vehicle passed all filters - add to display
            row_dict = vehicle.to_row_dict()
            values = [row_dict.get(col, "") for col in self.visible_columns]
            
            # Insert into treeview with color coding
            iid = self.tree.insert("", tk.END, values=values)
            self._apply_row_color_coding(iid, vehicle)
            
            # Map iid to vehicle object
            self.data_map[iid] = vehicle
            filtered_count += 1
        
        # Update summary to show filtered results
        self._update_summary_filtered(filtered_count)
    
    def _apply_row_color_coding(self, iid: str, vehicle: FleetVehicle) -> None:
        """
        Apply color coding to a row based on vehicle data quality.
        
        Args:
            iid: Treeview item ID
            vehicle: FleetVehicle object
        """
        # Determine colors based on processing success and data quality
        if not vehicle.processing_success:
            # Failed processing - red background
            self.tree.item(iid, tags=("failed",))
        elif vehicle.data_quality_score >= 80:
            # High quality - light green background
            self.tree.item(iid, tags=("high_quality",))
        elif vehicle.data_quality_score >= 50:
            # Medium quality - light yellow background
            self.tree.item(iid, tags=("medium_quality",))
        else:
            # Low quality - light orange background
            self.tree.item(iid, tags=("low_quality",))
    
    def _update_summary_filtered(self, filtered_count: int) -> None:
        """
        Update the summary display for filtered results.

        When a filter is active, shows stats (success/fail counts, avg MPG,
        unique makes) computed from the *filtered* subset — not the full fleet.

        Args:
            filtered_count: Number of items after filtering
        """
        if not self.data:
            self.summary_label.config(text="No data loaded")
            return

        total_count = len(self.data)

        if filtered_count == total_count:
            # No filtering applied — use normal full-fleet summary
            self._update_summary()
            return

        # Compute stats from the filtered vehicles only
        filtered_vehicles = list(self.data_map.values())
        successful = [v for v in filtered_vehicles if v.processing_success]
        failed_count = filtered_count - len(successful)

        mpg_values = [v.fuel_economy.combined_mpg for v in successful
                      if v.fuel_economy.combined_mpg > 0]
        avg_mpg = sum(mpg_values) / len(mpg_values) if mpg_values else 0

        unique_makes = len({v.vehicle_id.make for v in successful
                           if v.vehicle_id.make})

        # Build summary parts
        parts = [f"Showing {filtered_count} of {total_count}"]
        if failed_count > 0:
            parts.append(f"✗ {failed_count} failed")
        if unique_makes > 0:
            parts.append(f"{unique_makes} make{'s' if unique_makes != 1 else ''}")
        if avg_mpg > 0:
            parts.append(f"Avg MPG: {avg_mpg:.1f}")

        text_color = "#4682B4"  # Steel blue for filtered results
        self.summary_label.config(text="  ·  ".join(parts), foreground=text_color)

        # MPG coverage badge (same bucketing as full-fleet summary)
        real_mpg = sum(1 for v in successful
                       if v.fuel_economy.combined_mpg > 0 and not v.fuel_economy.mpg_is_estimate)
        est_mpg = sum(1 for v in successful
                      if v.fuel_economy.combined_mpg > 0 and v.fuel_economy.mpg_is_estimate)
        missing_mpg = len(successful) - real_mpg - est_mpg

        if missing_mpg > 0:
            badge = f"⚠ {missing_mpg}/{len(successful)} missing MPG"
            if est_mpg > 0:
                badge += f"  ·  {est_mpg} estimated"
            self.count_label.config(text=badge, foreground="#CC5500")
        elif est_mpg > 0:
            self.count_label.config(
                text=f"{real_mpg}/{len(successful)} real MPG  ·  {est_mpg} estimated",
                foreground="#8B6914"
            )
        else:
            self.count_label.config(text=f"filtered from {total_count}",
                                    foreground=Colors.TEXT_TERTIARY)
    
    def _clear_filter(self):
        """Clear search filter and quick filters."""
        self.search_var.set("")
        self.status_filter_var.set("all")
        self.quality_filter_var.set("all")
        self.populate_data()
    
    def _export_dialog(self):
        """Unified export: CSV or Excel via a single save dialog."""
        from tkinter import filedialog
        import csv
        import datetime

        # Get currently displayed data (respects active filters/sorting)
        displayed_vehicles = []
        for iid in self.tree.get_children():
            if iid in self.data_map:
                displayed_vehicles.append(self.data_map[iid])

        if not displayed_vehicles:
            messagebox.showwarning("No Data", "No data to export.")
            return

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        filepath = filedialog.asksaveasfilename(
            title="Export Fleet Data",
            initialfile=f"fleet_analysis_{timestamp}",
            defaultextension=".csv",
            filetypes=[
                ("CSV (Excel compatible)", "*.csv"),
                ("Excel Workbook", "*.xlsx"),
            ]
        )

        if not filepath:
            return

        try:
            if filepath.lower().endswith('.xlsx'):
                self._write_xlsx(filepath, displayed_vehicles)
            else:
                self._write_csv(filepath, displayed_vehicles)

            messagebox.showinfo(
                "Export Complete",
                f"Exported {len(displayed_vehicles)} vehicles to:\n{os.path.basename(filepath)}"
            )
        except Exception as e:
            logger.error(f"Export failed: {e}")
            messagebox.showerror("Export Failed", f"Error exporting data:\n{e}")

    def _get_export_columns(self, displayed_vehicles):
        """Build ordered column list for export."""
        # Core columns in a logical order
        export_cols = [
            "VIN", "Processing Status", "Data Quality", "Year", "Make", "Model",
            "FuelTypePrimary", "BodyClass", "GVWR",
            "MPG Combined", "MPG City", "MPG Highway",
            "CO2 emissions", "Asset ID", "Department", "Location",
            "Odometer", "Annual Mileage", "Processing Error"
        ]
        # Append any custom fields not already listed
        if displayed_vehicles:
            sample = displayed_vehicles[0].to_row_dict()
            skip = {"co2A", "rangeA", "Assumed Vehicle (Text)", "Assumed Vehicle (ID)"}
            for key in sample:
                if key not in export_cols and key not in skip:
                    export_cols.append(key)
        return export_cols

    def _write_csv(self, filepath, vehicles):
        """Write vehicles to CSV with UTF-8 BOM for Excel compatibility."""
        import csv
        columns = self._get_export_columns(vehicles)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for v in vehicles:
                row = v.to_row_dict()
                writer.writerow([row.get(c, "") for c in columns])

    def _write_xlsx(self, filepath, vehicles):
        """Write vehicles to an Excel workbook (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "Excel export requires the openpyxl package.\n\n"
                "Install it with: pip install openpyxl\n\n"
                "Alternatively, export as CSV."
            )
            return

        columns = self._get_export_columns(vehicles)
        wb = Workbook()
        ws = wb.active
        ws.title = "Fleet Analysis"

        # Header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, v in enumerate(vehicles, 2):
            row = v.to_row_dict()
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        # Auto-width columns (cap at 40)
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row_idx in range(2, min(len(vehicles) + 2, 52)):  # sample 50 rows
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

        wb.save(filepath)
    
    def _on_selection_change(self, event):
        """Handle selection change in treeview."""
        # Get selected items
        selected_iids = self.tree.selection()
        
        # Map to vehicle objects
        selected_vehicles = [self.data_map[iid] for iid in selected_iids if iid in self.data_map]
        
        # Call callback if provided
        if self.on_selection_change_callback:
            self.on_selection_change_callback(selected_vehicles)
    
    def _show_details(self, event=None):
        """Show details for the selected vehicle."""
        # Get selected item
        selected_iids = self.tree.selection()
        
        if not selected_iids:
            # If called from menu without selection, just return
            return
        
        # Get the first selected vehicle
        if selected_iids[0] in self.data_map:
            vehicle = self.data_map[selected_iids[0]]
            
            # Create details window
            self._create_details_window(vehicle)
    
    def _create_details_window(self, vehicle):
        """
        Create a details window for the specified vehicle.
        
        Args:
            vehicle: FleetVehicle object to display
        """
        # Create a new toplevel window
        details_window = tk.Toplevel(self)
        details_window.title(f"Vehicle Details: {vehicle.vin}")
        details_window.geometry("600x500")
        details_window.transient(self.master)  # Set as transient to main window
        
        # Make the window modal
        details_window.grab_set()
        
        # Create a notebook with tabs for different categories
        notebook = ttk.Notebook(details_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Overview tab
        overview_frame = ttk.Frame(notebook)
        notebook.add(overview_frame, text="Overview")
        
        # Create a grid of labels for overview data
        row = 0
        
        # Add VIN row
        ttk.Label(overview_frame, text="VIN:", font=("", 10, "bold")).grid(
            row=row, column=0, sticky=tk.W, padx=5, pady=2
        )
        ttk.Label(overview_frame, text=vehicle.vin).grid(
            row=row, column=1, sticky=tk.W, padx=5, pady=2
        )
        row += 1
        
        # Basic vehicle info section
        ttk.Separator(overview_frame, orient=tk.HORIZONTAL).grid(
            row=row, column=0, columnspan=2, sticky=tk.EW, pady=5
        )
        row += 1
        
        ttk.Label(overview_frame, text="Vehicle Information", font=("", 10, "bold", "underline")).grid(
            row=row, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2
        )
        row += 1
        
        # Add basic vehicle info
        basic_info = [
            ("Year:", vehicle.vehicle_id.year),
            ("Make:", vehicle.vehicle_id.make),
            ("Model:", vehicle.vehicle_id.model),
            ("Fuel Type:", vehicle.vehicle_id.fuel_type),
            ("Body Class:", vehicle.vehicle_id.body_class),
            ("GVWR:", vehicle.vehicle_id.gvwr)
        ]
        
        for label, value in basic_info:
            ttk.Label(overview_frame, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=5, pady=2
            )
            ttk.Label(overview_frame, text=value).grid(
                row=row, column=1, sticky=tk.W, padx=5, pady=2
            )
            row += 1
        
        # Fuel economy section
        ttk.Separator(overview_frame, orient=tk.HORIZONTAL).grid(
            row=row, column=0, columnspan=2, sticky=tk.EW, pady=5
        )
        row += 1
        
        ttk.Label(overview_frame, text="Fuel Economy", font=("", 10, "bold", "underline")).grid(
            row=row, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2
        )
        row += 1
        
        # Add fuel economy info
        economy_info = [
            ("City MPG:", vehicle.fuel_economy.city_mpg),
            ("Highway MPG:", vehicle.fuel_economy.highway_mpg),
            ("Combined MPG:", vehicle.fuel_economy.combined_mpg),
            ("CO2 Emissions:", f"{vehicle.fuel_economy.co2_primary} g/mile"),
            ("Annual Fuel Cost:", f"${vehicle.fuel_economy.fuel_cost_primary}")
        ]
        
        for label, value in economy_info:
            ttk.Label(overview_frame, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=5, pady=2
            )
            ttk.Label(overview_frame, text=value).grid(
                row=row, column=1, sticky=tk.W, padx=5, pady=2
            )
            row += 1
        
        # Fleet data section
        ttk.Separator(overview_frame, orient=tk.HORIZONTAL).grid(
            row=row, column=0, columnspan=2, sticky=tk.EW, pady=5
        )
        row += 1
        
        ttk.Label(overview_frame, text="Fleet Data", font=("", 10, "bold", "underline")).grid(
            row=row, column=0, columnspan=2, sticky=tk.W, padx=5, pady=2
        )
        row += 1
        
        # Add fleet data
        fleet_info = [
            ("Asset ID:", vehicle.asset_id or "N/A"),
            ("Department:", vehicle.department or "N/A"),
            ("Location:", vehicle.location or "N/A"),
            ("Odometer:", f"{vehicle.odometer:,.0f} miles" if vehicle.odometer else "N/A"),
            ("Annual Mileage:", f"{vehicle.annual_mileage:,.0f} miles" if vehicle.annual_mileage else "N/A")
        ]
        
        for label, value in fleet_info:
            ttk.Label(overview_frame, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=5, pady=2
            )
            ttk.Label(overview_frame, text=value).grid(
                row=row, column=1, sticky=tk.W, padx=5, pady=2
            )
            row += 1
        
        # Technical details tab
        tech_frame = ttk.Frame(notebook)
        notebook.add(tech_frame, text="Technical Details")
        
        # Create a scrollable text widget for all technical details
        tech_text = tk.Text(
            tech_frame,
            wrap=tk.WORD,
            width=80,
            height=20,
            bg=PRIMARY_HEX_2,
            state=tk.DISABLED
        )
        tech_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Add scrollbar
        tech_scrollbar = ttk.Scrollbar(tech_frame, command=tech_text.yview)
        tech_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tech_text.config(yscrollcommand=tech_scrollbar.set)
        
        # Populate technical details
        tech_text.config(state=tk.NORMAL)
        
        # Vehicle ID details
        tech_text.insert(tk.END, "Vehicle Identification\n", "heading")
        tech_text.insert(tk.END, "-" * 50 + "\n", "separator")
        
        vehicle_id_details = [
            ("Engine Displacement:", vehicle.vehicle_id.engine_displacement),
            ("Engine Cylinders:", vehicle.vehicle_id.engine_cylinders),
            ("Drive Type:", vehicle.vehicle_id.drive_type),
            ("Transmission:", vehicle.vehicle_id.transmission)
        ]
        
        for label, value in vehicle_id_details:
            tech_text.insert(tk.END, f"{label} {value}\n", "detail")
        
        tech_text.insert(tk.END, "\n")
        
        # Fuel Economy details
        tech_text.insert(tk.END, "Fuel Economy Details\n", "heading")
        tech_text.insert(tk.END, "-" * 50 + "\n", "separator")
        
        # Get all raw data
        for key, value in vehicle.fuel_economy.raw_data.items():
            if value:  # Only show non-empty values
                tech_text.insert(tk.END, f"{key}: {value}\n", "detail")
        
        tech_text.insert(tk.END, "\n")
        
        # Custom fields
        if vehicle.custom_fields:
            tech_text.insert(tk.END, "Custom Fields\n", "heading")
            tech_text.insert(tk.END, "-" * 50 + "\n", "separator")
            
            for key, value in vehicle.custom_fields.items():
                tech_text.insert(tk.END, f"{key}: {value}\n", "detail")
            
            tech_text.insert(tk.END, "\n")
        
        # Configure text styles
        tech_text.tag_configure("heading", font=("", 12, "bold"), foreground=PRIMARY_HEX_1)
        tech_text.tag_configure("separator", foreground=PRIMARY_HEX_3)
        tech_text.tag_configure("detail", font=("", 10, ""))
        
        tech_text.config(state=tk.DISABLED)
        
        # Raw Data tab
        raw_frame = ttk.Frame(notebook)
        notebook.add(raw_frame, text="Raw Data")
        
        # Create a scrollable text widget for all raw data
        raw_text = tk.Text(
            raw_frame,
            wrap=tk.WORD,
            width=80,
            height=20,
            bg=PRIMARY_HEX_2,
            state=tk.DISABLED
        )
        raw_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Add scrollbar
        raw_scrollbar = ttk.Scrollbar(raw_frame, command=raw_text.yview)
        raw_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        raw_text.config(yscrollcommand=raw_scrollbar.set)
        
        # Populate raw data
        raw_text.config(state=tk.NORMAL)
        
        # Get the full data dictionary
        import json
        raw_data = json.dumps(vehicle.to_dict(), indent=2)
        raw_text.insert(tk.END, raw_data)
        
        raw_text.config(state=tk.DISABLED)
        
        # Close button
        close_button = ttk.Button(
            details_window,
            text="Close",
            command=details_window.destroy
        )
        close_button.pack(side=tk.BOTTOM, pady=10)
    
    def _show_context_menu(self, event):
        """Show the context menu for the treeview."""
        # Check if mouse was clicked on an item
        item = self.tree.identify_row(event.y)
        if item:
            # Select the item under the mouse if not already selected
            if item not in self.tree.selection():
                self.tree.selection_set(item)
            
            # Show the context menu
            self.context_menu.tk_popup(event.x_root, event.y_root)
    
    def _analyze_selected(self):
        """Analyze the selected vehicles."""
        # Get selected vehicles
        selected_iids = self.tree.selection()
        selected_vehicles = [self.data_map[iid] for iid in selected_iids if iid in self.data_map]
        
        if not selected_vehicles:
            messagebox.showinfo("No Selection", "Please select one or more vehicles to analyze.")
            return
        
        # Create an analysis popup
        analysis_window = tk.Toplevel(self)
        analysis_window.title(f"Analysis of {len(selected_vehicles)} Vehicles")
        analysis_window.geometry("600x500")
        analysis_window.transient(self.master)  # Set as transient to main window
        
        # Make the window modal
        analysis_window.grab_set()
        
        # Create a notebook with tabs for different analyses
        notebook = ttk.Notebook(analysis_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Summary tab
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text="Summary")
        
        # Calculate summary statistics
        mpg_values = [v.fuel_economy.combined_mpg for v in selected_vehicles if v.fuel_economy.combined_mpg > 0]
        co2_values = [v.fuel_economy.co2_primary for v in selected_vehicles if v.fuel_economy.co2_primary > 0]
        
        avg_mpg = sum(mpg_values) / len(mpg_values) if mpg_values else 0
        avg_co2 = sum(co2_values) / len(co2_values) if co2_values else 0
        
        # Create summary text
        summary_text = f"""
Summary of Selected Vehicles ({len(selected_vehicles)})

• Average Combined MPG: {avg_mpg:.1f}
• Average CO2 Emissions: {avg_co2:.1f} g/mile
        """
        
        # Add summary label
        ttk.Label(
            summary_frame,
            text=summary_text,
            font=("", 12, ""),
            wraplength=550,
            justify=tk.LEFT
        ).pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # TODO: Add more analysis tabs as needed
        
        # Close button
        close_button = ttk.Button(
            analysis_window,
            text="Close",
            command=analysis_window.destroy
        )
        close_button.pack(side=tk.BOTTOM, pady=10)
    
    def _select_all(self):
        """Select all items in the treeview."""
        self.tree.selection_set(self.tree.get_children())
    
    def _deselect_all(self):
        """Deselect all items in the treeview."""
        self.tree.selection_remove(self.tree.selection())

    # -------------------------------------------------------------------------
    # Database integration
    # -------------------------------------------------------------------------

    def _override_acf_category(self) -> None:
        """Context menu action: manually override the ACF category of the selected vehicle.

        Shows a small dialog with a dropdown of all valid ACF categories.
        If the user confirms, applies the override and fires on_acf_override_callback.
        """
        from ui.analysis_panel import ACF_LABELS, ACF_CODE_TO_LABEL, ACF_LABEL_TO_CODE, _show_acf_ev_year_dialog, AnalysisPanel
        from analysis.electrification_timeline import assign_electrification_years

        selected_iids = self.tree.selection()
        if not selected_iids:
            messagebox.showinfo("No Selection", "Select a vehicle row first.", parent=self)
            return

        vehicle = self.data_map.get(selected_iids[0])
        if vehicle is None:
            return

        old_label = vehicle.custom_fields.get("ACF Category", "")
        old_code = vehicle.custom_fields.get("_acf_code", "")

        # ── Category picker dialog ────────────────────────────────────────────
        picker = tk.Toplevel(self)
        picker.title("Override Vehicle ACF Designation")
        picker.resizable(False, False)
        picker.transient(self.winfo_toplevel())
        picker.grab_set()
        try:
            picker.geometry(
                f"+{self.winfo_rootx() + 80}+{self.winfo_rooty() + 80}"
            )
        except tk.TclError:
            pass

        vehicle_label = " ".join(filter(None, [
            str(vehicle.vehicle_id.year or ""),
            vehicle.vehicle_id.make or "",
            vehicle.vehicle_id.model or "",
        ])).strip() or "Unknown Vehicle"

        ttk.Label(
            picker,
            text=f"Override ACF Designation for:\n{vehicle_label}",
            font=("TkDefaultFont", 10, "bold"),
        ).pack(padx=20, pady=(16, 4), anchor="w")

        ttk.Label(
            picker,
            text=f"Current: {old_label}",
            foreground="#555",
        ).pack(padx=20, pady=(0, 8), anchor="w")

        ttk.Label(picker, text="New designation:").pack(padx=20, anchor="w")
        options = list(ACF_CODE_TO_LABEL.values())
        combo = ttk.Combobox(picker, values=options, state="readonly", width=48)
        current_option = old_label if old_label in options else options[0]
        combo.set(current_option)
        combo.pack(padx=20, pady=(4, 12), anchor="w")

        chosen = tk.StringVar(value="cancel")

        def on_ok():
            chosen.set(combo.get())
            picker.destroy()

        def on_cancel():
            picker.destroy()

        btn_row = ttk.Frame(picker)
        btn_row.pack(padx=20, pady=(0, 16), anchor="e")
        ttk.Button(btn_row, text="OK", style="Primary.TButton",
                   command=on_ok).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_row, text="Cancel",
                   command=on_cancel).pack(side=tk.LEFT)

        picker.wait_window()
        raw_choice = chosen.get()
        if raw_choice == "cancel" or not raw_choice:
            return

        new_code = ACF_LABEL_TO_CODE.get(raw_choice, raw_choice)
        if new_code == old_code:
            return  # No change

        # ── EV year warning dialog ────────────────────────────────────────────
        current_ev_year = vehicle.custom_fields.get("Proposed EV Year", "N/A")
        ev_choice = _show_acf_ev_year_dialog(
            self, old_code, new_code, current_ev_year
        )
        if ev_choice == "cancel":
            return

        # Apply the ACF override
        AnalysisPanel._apply_acf_override(vehicle, new_code)

        if ev_choice == "recalculate":
            vehicle.custom_fields.pop("EV Year Overridden", None)
            vehicle.custom_fields.pop("System Recommended EV Year", None)
            assign_electrification_years([vehicle], fleet_type=self._fleet_type)

        # Refresh the Results table row
        self.refresh()

        # Notify main window so other panels (Timeline, Analysis) can sync
        if self.on_acf_override_callback:
            self.on_acf_override_callback(vehicle)

    def _save_mpg_to_database(self) -> None:
        """
        Context menu action: save the selected vehicle's MPG to the reference DB.
        Opens a confirmation/entry dialog pre-filled with the vehicle's data.
        """
        if self.db_manager is None:
            messagebox.showwarning(
                "Database Unavailable",
                "The vehicle reference database could not be opened.\n"
                "Check the application log for details.",
                parent=self,
            )
            return

        selected_iids = self.tree.selection()
        if not selected_iids:
            messagebox.showinfo("No Selection", "Select a vehicle row first.", parent=self)
            return

        vehicle = self.data_map.get(selected_iids[0])
        if vehicle is None:
            return

        self._show_save_mpg_dialog(vehicle)

    def _show_save_mpg_dialog(self, vehicle) -> None:
        """
        Show a dialog for saving vehicle MPG data to the reference database.

        Pre-fills fields if the vehicle already has MPG data.
        Requires manual entry if MPG is zero or missing.
        """
        has_mpg = bool(vehicle.fuel_economy.combined_mpg and
                       vehicle.fuel_economy.combined_mpg > 0)

        dialog = tk.Toplevel(self)
        dialog.title("Save MPG to Database")
        dialog.geometry("440x460")
        dialog.resizable(False, False)
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()

        # Header
        vehicle_label = " ".join(filter(None, [
            str(vehicle.vehicle_id.year  or ""),
            vehicle.vehicle_id.make  or "",
            vehicle.vehicle_id.model or "",
        ])).strip() or "Unknown Vehicle"

        ttk.Label(
            dialog,
            text=f"Save MPG for: {vehicle_label}",
            font=(Fonts.FAMILY_SANS, Fonts.SIZE_H3, "bold"),
        ).pack(padx=16, pady=(14, 2), anchor="w")

        if not has_mpg:
            ttk.Label(
                dialog,
                text="No MPG data found for this vehicle. Enter values manually:",
                foreground=Colors.WARNING,
            ).pack(padx=16, pady=(0, 6), anchor="w")
        else:
            ttk.Label(
                dialog,
                text="Review the pre-filled data and save to the database.",
                foreground=Colors.TEXT_SECONDARY,
            ).pack(padx=16, pady=(0, 6), anchor="w")

        ttk.Separator(dialog, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=16, pady=(0, 8))

        # Form
        form = ttk.Frame(dialog)
        form.pack(fill=tk.X, padx=16)
        form.columnconfigure(1, weight=1)

        fields: Dict[str, tk.StringVar] = {}

        def add_row(label: str, key: str, default: str = "", width: int = 22) -> None:
            row = form.grid_size()[1]
            ttk.Label(form, text=label, anchor="w", width=16).grid(
                row=row, column=0, sticky="w", pady=2
            )
            var = tk.StringVar(value=default)
            fields[key] = var
            ttk.Entry(form, textvariable=var, width=width).grid(
                row=row, column=1, sticky="ew", pady=2, padx=(0, 4)
            )

        def add_combo(label: str, key: str, values: list, default: str = "",
                      width: int = 20) -> None:
            row = form.grid_size()[1]
            ttk.Label(form, text=label, anchor="w", width=16).grid(
                row=row, column=0, sticky="w", pady=2
            )
            var = tk.StringVar(value=default)
            fields[key] = var
            ttk.Combobox(form, textvariable=var, values=values, width=width).grid(
                row=row, column=1, sticky="ew", pady=2, padx=(0, 4)
            )

        year_val  = str(vehicle.vehicle_id.year  or "")
        make_val  = vehicle.vehicle_id.make  or ""
        model_val = vehicle.vehicle_id.model or ""
        fuel_val  = (vehicle.vehicle_id.fuel_type or "").lower()

        add_row("Year:",     "year",  year_val,  8)
        add_row("Make *:",   "make",  make_val,  22)
        add_row("Model *:",  "model", model_val, 22)
        add_combo("Fuel Type:", "fuel_type",
                  ["", "gasoline", "diesel", "flex", "hybrid", "cng", "propane"],
                  fuel_val, 18)

        mpg_c   = str(int(vehicle.fuel_economy.combined_mpg)
                      if has_mpg and vehicle.fuel_economy.combined_mpg == int(vehicle.fuel_economy.combined_mpg)
                      else round(vehicle.fuel_economy.combined_mpg, 1)) if has_mpg else ""
        mpg_city = str(int(vehicle.fuel_economy.city_mpg)
                       if has_mpg and vehicle.fuel_economy.city_mpg and vehicle.fuel_economy.city_mpg == int(vehicle.fuel_economy.city_mpg)
                       else (round(vehicle.fuel_economy.city_mpg, 1) if has_mpg and vehicle.fuel_economy.city_mpg else "")) if has_mpg else ""
        mpg_hwy  = str(int(vehicle.fuel_economy.highway_mpg)
                       if has_mpg and vehicle.fuel_economy.highway_mpg and vehicle.fuel_economy.highway_mpg == int(vehicle.fuel_economy.highway_mpg)
                       else (round(vehicle.fuel_economy.highway_mpg, 1) if has_mpg and vehicle.fuel_economy.highway_mpg else "")) if has_mpg else ""

        add_row("MPG Combined *:", "mpg_combined", mpg_c,    8)
        add_row("MPG City:",       "mpg_city",     mpg_city, 8)
        add_row("MPG Highway:",    "mpg_highway",  mpg_hwy,  8)

        # GVWR weight class — pre-select from decoded vehicle data
        _GVWR_OPTIONS = [
            "",
            "Class 2b–4 (8,501–19,500 lbs)",
            "Class 5–8a (19,501–33,000 lbs)",
            "Class 8b (33,001+ lbs)",
        ]
        _GVWR_BOUNDS = {
            "Class 2b–4 (8,501–19,500 lbs)":   (8501,  19500),
            "Class 5–8a (19,501–33,000 lbs)":  (19501, 33000),
            "Class 8b (33,001+ lbs)":                    (33001, None),
        }
        gvwr_lbs = vehicle.vehicle_id.gvwr_pounds or 0
        if gvwr_lbs > 33000:
            _gvwr_default = "Class 8b (33,001+ lbs)"
        elif gvwr_lbs > 19500:
            _gvwr_default = "Class 5–8a (19,501–33,000 lbs)"
        elif gvwr_lbs > 8500:
            _gvwr_default = "Class 2b–4 (8,501–19,500 lbs)"
        else:
            _gvwr_default = ""
        add_combo("GVWR Class:", "gvwr_class", _GVWR_OPTIONS, _gvwr_default, 28)

        add_combo("Source:", "source",
                  ["analyst", "manufacturer_spec", "fuelly", "epa_label", "fleet_record"],
                  "analyst", 18)

        # Notes
        notes_row = form.grid_size()[1]
        ttk.Label(form, text="Notes:", anchor="nw", width=16).grid(
            row=notes_row, column=0, sticky="nw", pady=2
        )
        notes_text = tk.Text(form, height=3, width=28, wrap=tk.WORD)
        notes_text.grid(row=notes_row, column=1, sticky="ew", pady=2, padx=(0, 4))

        ttk.Separator(dialog, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=16, pady=(10, 4))

        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=16, pady=(0, 12))

        def do_save() -> None:
            make  = fields["make"].get().strip()
            model = fields["model"].get().strip()
            mpg_c_str = fields["mpg_combined"].get().strip()

            if not make:
                messagebox.showerror("Validation", "Make is required.", parent=dialog)
                return
            if not model:
                messagebox.showerror("Validation", "Model is required.", parent=dialog)
                return
            if not mpg_c_str:
                messagebox.showerror("Validation", "MPG Combined is required.", parent=dialog)
                return
            try:
                mpg_combined = float(mpg_c_str)
                if mpg_combined <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror(
                    "Validation",
                    "MPG Combined must be a number greater than zero.",
                    parent=dialog
                )
                return

            year_str = fields["year"].get().strip()
            try:
                year = int(year_str) if year_str else None
            except ValueError:
                messagebox.showerror("Validation", "Year must be a number.", parent=dialog)
                return

            def _safe_float(key: str) -> float:
                try:
                    return float(fields[key].get().strip() or 0)
                except ValueError:
                    return 0.0

            gvwr_sel = fields["gvwr_class"].get().strip()
            gvwr_min, gvwr_max = _GVWR_BOUNDS.get(gvwr_sel, (None, None))

            try:
                self.db_manager.add_ice_vehicle(
                    year         = year,
                    make         = make,
                    model        = model,
                    fuel_type    = fields["fuel_type"].get().strip() or None,
                    body_class   = vehicle.vehicle_id.body_class or None,
                    gvwr_lbs_min = gvwr_min,
                    gvwr_lbs_max = gvwr_max,
                    mpg_combined = mpg_combined,
                    mpg_city     = _safe_float("mpg_city"),
                    mpg_highway  = _safe_float("mpg_highway"),
                    notes        = notes_text.get("1.0", tk.END).strip(),
                    source       = fields["source"].get().strip() or "analyst",
                )
                messagebox.showinfo(
                    "Saved",
                    f"MPG data saved for {make} {model}.\n"
                    "Future processing runs will use this value.",
                    parent=dialog,
                )
                dialog.destroy()
            except Exception as e:
                logger.error(f"Failed to save MPG to database: {e}")
                messagebox.showerror("Save Failed", f"An error occurred:\n{e}", parent=dialog)

        ttk.Button(
            btn_frame, text="Cancel",
            command=dialog.destroy
        ).pack(side=tk.RIGHT, padx=(4, 0))

        ttk.Button(
            btn_frame, text="Save to Database",
            command=do_save,
            style="Primary.TButton"
        ).pack(side=tk.RIGHT)

    def select_by_vin(self, vin: str) -> bool:
        """Select the row matching *vin*, scroll it into view, and return True if found."""
        target = vin.upper().strip()
        for iid, vehicle in self.data_map.items():
            if vehicle.vin and vehicle.vin.upper() == target:
                self.tree.selection_set(iid)
                self.tree.see(iid)
                self.tree.focus(iid)
                return True
        return False

    def set_data(self, data):
        """
        Set the data for the results panel.

        Args:
            data: List of FleetVehicle objects
        """
        self.data = data or []
        self.populate_data()
        self._update_summary()

        if self.on_selection_change_callback:
            self.on_selection_change_callback([])
    
    def _update_summary(self):
        """Update the processing summary display with fleet statistics."""
        if not self.data:
            self.summary_label.config(text="No data loaded", foreground="#718096")
            self.count_label.config(text="")
            return

        total_count = len(self.data)
        successful_count = sum(1 for v in self.data if v.processing_success)
        failed_count = total_count - successful_count

        # Gather fleet stats for successful vehicles
        mpg_values = [v.fuel_economy.combined_mpg for v in self.data
                      if v.processing_success and v.fuel_economy.combined_mpg > 0]
        avg_mpg = sum(mpg_values) / len(mpg_values) if mpg_values else 0

        unique_makes = len({v.vehicle_id.make for v in self.data
                           if v.processing_success and v.vehicle_id.make})

        # Build summary parts
        parts = []
        if failed_count == 0:
            parts.append(f"✓ {successful_count} vehicles")
            text_color = "#2E8B57"
        elif successful_count == 0:
            parts.append(f"✗ {failed_count} failed")
            text_color = "#DC143C"
        else:
            parts.append(f"✓ {successful_count} successful · ✗ {failed_count} failed")
            text_color = "#FF8C00"

        if unique_makes > 0:
            parts.append(f"{unique_makes} make{'s' if unique_makes != 1 else ''}")
        if avg_mpg > 0:
            parts.append(f"Avg MPG: {avg_mpg:.1f}")

        self.summary_label.config(text="  ·  ".join(parts), foreground=text_color)

        # Right-side count badge — highlights MPG coverage gaps prominently.
        # Buckets: real MPG (API/scraper), estimated MPG (EPA class average), missing MPG (0).
        successful_vehicles = [v for v in self.data if v.processing_success]
        real_mpg_count = sum(
            1 for v in successful_vehicles
            if v.fuel_economy.combined_mpg > 0 and not v.fuel_economy.mpg_is_estimate
        )
        est_mpg_count = sum(
            1 for v in successful_vehicles
            if v.fuel_economy.combined_mpg > 0 and v.fuel_economy.mpg_is_estimate
        )
        missing_mpg_count = successful_count - real_mpg_count - est_mpg_count

        if missing_mpg_count > 0:
            badge_text = f"⚠ {missing_mpg_count} of {successful_count} missing MPG"
            if est_mpg_count > 0:
                badge_text += f"  ·  {est_mpg_count} estimated"
            self.count_label.config(text=badge_text, foreground="#CC5500")
        elif est_mpg_count > 0:
            self.count_label.config(
                text=f"{real_mpg_count}/{successful_count} real MPG  ·  {est_mpg_count} estimated",
                foreground="#8B6914"
            )
        else:
            self.count_label.config(
                text=f"{total_count} total",
                foreground=Colors.TEXT_TERTIARY
            )
    
    def get_data(self):
        """
        Get the current data.
        
        Returns:
            List of FleetVehicle objects
        """
        return self.data
    
    def get_selected_data(self):
        """
        Get the currently selected data.
        
        Returns:
            List of selected FleetVehicle objects
        """
        selected_iids = self.tree.selection()
        return [self.data_map[iid] for iid in selected_iids if iid in self.data_map]
    
    def set_visible_columns(self, columns):
        """
        Set the visible columns.
        
        Args:
            columns: List of column IDs to display
        """
        # Validate columns
        valid_columns = [col for col in columns if col in self.all_columns_map]
        
        if not valid_columns:
            # Ensure at least some columns are visible
            valid_columns = DEFAULT_VISIBLE_COLUMNS
        
        self.visible_columns = valid_columns
        
        # Update the treeview
        self._update_columns()
        
        # Call callback
        if self.on_column_change_callback:
            self.on_column_change_callback(self.visible_columns)
    
    def get_visible_columns(self):
        """
        Get the currently visible columns.
        
        Returns:
            List of visible column IDs
        """
        return self.visible_columns
    
    def get_all_columns(self):
        """
        Get all available columns with their display names.
        
        Returns:
            List of (column_id, display_name) tuples
        """
        return [(col_id, display_name) for col_id, display_name in self.all_columns_map.items()]
    
    def _update_columns(self):
        """Update the treeview columns."""
        # Remember the current data
        current_data = self.data_map.values()
        
        # Clear the treeview
        self.tree.delete(*self.tree.get_children())
        
        # Configure new columns
        self.tree.config(columns=self.visible_columns)
        
        # Create new headers with content-aware widths
        for col in self.visible_columns:
            display_name = self.all_columns_map.get(col, col)
            self.tree.heading(col, text=display_name, command=lambda c=col: self._sort_by_column(c))

            width, anchor = self._get_column_width_and_anchor(col, display_name)
            self.tree.column(col, width=width, minwidth=50, anchor=anchor)
        
        # Repopulate data
        self.populate_data()
    
    def populate_data(self):
        """Populate the treeview with data from self.data."""
        # Clear existing data
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.data_map = {}

        if not self.data:
            return

        self._build_column_map()

        # Add data to treeview
        for vehicle in self.data:
            try:
                row_data = []
                for col_id in self.visible_columns:
                    if col_id in self.all_columns_map:
                        value = self._get_vehicle_value(vehicle, col_id)

                        # Special formatting for failed vehicles
                        if not vehicle.processing_success and col_id == "processing_error":
                            value = f"❌ {value}" if value else "❌ Unknown Error"
                        elif not vehicle.processing_success and col_id == "vin":
                            value = f"⚠️ {value}" if value else "⚠️ No VIN"

                        row_data.append(str(value) if value is not None else "")
                    else:
                        row_data.append("")

                iid = self.tree.insert("", tk.END, values=row_data)
                self.data_map[iid] = vehicle
                self._apply_row_color_coding(iid, vehicle)

            except Exception as ve:
                logger.error(f"Error adding vehicle to results table: {ve}")
    
    def refresh(self):
        """Refresh the display."""
        # Clear any filter
        self.search_var.set("")
        
        # Repopulate data
        self.populate_data()
    
    def copy_selection(self):
        """Copy selected rows to clipboard in tab-delimited format."""
        # Get selected items
        selection = self.tree.selection()
        
        if not selection:
            return
        
        # Prepare clipboard text
        clipboard_lines = []
        
        # Add header row
        headers = [self.all_columns_map.get(col, col) for col in self.visible_columns]
        clipboard_lines.append("\t".join(headers))
        
        # Add selected rows
        for iid in selection:
            values = [self.tree.set(iid, col) for col in self.visible_columns]
            clipboard_lines.append("\t".join(values))
        
        # Join lines and copy to clipboard
        clipboard_text = "\n".join(clipboard_lines)
        self.clipboard_clear()
        self.clipboard_append(clipboard_text)
    
    def on_resize(self):
        """Handle resize events."""
        # Adjust column widths if needed
        pass

    def _get_vehicle_value(self, vehicle: FleetVehicle, col_id: str) -> Any:
        """
        Get a field value from a vehicle object by column ID.
        
        Args:
            vehicle: FleetVehicle object
            col_id: Column identifier
            
        Returns:
            Field value or empty string if not found
        """
        # Get the row dictionary which has all the properly formatted data
        row_dict = vehicle.to_row_dict()
        
        # Try to get the value directly by column ID
        if col_id in row_dict:
            return row_dict[col_id]
        
        # If not found, try some common mappings
        column_mappings = {
            "vin": "VIN",
            "year": "Year", 
            "make": "Make",
            "model": "Model",
            "fuel_type": "FuelTypePrimary",
            "body_class": "BodyClass",
            "mpg_combined": "MPG Combined",
            "co2_emissions": "CO2 emissions",
            "commercial_category": "Commercial Category",
            "gvwr_lbs": "GVWR (lbs)",
            "data_quality": "Data Quality",
            "processing_status": "Processing Status",
            "annual_mileage": "Annual Mileage",
            "asset_id": "Asset ID",
            "department": "Department",
            "location": "Location",
            "odometer": "Odometer"
        }
        
        # Try the mapping
        mapped_key = column_mappings.get(col_id.lower(), col_id)
        if mapped_key in row_dict:
            return row_dict[mapped_key]
        
        # Still not found, try the vehicle's get_field method as last resort
        try:
            return vehicle.get_field(col_id)
        except Exception:
            return ""